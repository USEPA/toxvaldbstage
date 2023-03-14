# -*- coding: utf-8 -*-
"""
Created on Tue Jan 24 16:03:18 2023

@author: JWALL01
"""

from bs4 import BeautifulSoup
import requests
import pandas as pd
import time
import openpyxl
from tqdm import tqdm
from os import listdir, makedirs, path

# Prep output folder
if not path.isdir("excel_sheets"):
    makedirs("excel_sheets")
# Get previously downloaded chemical list
chemicals_downloaded = [
    filename.split("_")[2]
    for filename in listdir("excel_sheets")
    # Only pull XLSX files
    if filename.endswith(".xlsx")
]

# Read in data and grab the list of URLs to scrape
# Pandas does not support hyperlinks, so we need to use openpyxl
cphea = openpyxl.load_workbook("pprtv_cphea_chemicals_20230123.xlsx")
sheet = cphea["pprtv_cphea_chemicals_20230123"]
url_list = [
    sheet.cell(row=i, column=2).hyperlink.target for i in range(2, sheet.max_row + 1)
]

chemical_list = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]

# Dataframe of chemical information
in_chemicals = pd.DataFrame({"chems": chemical_list, "url": url_list})
# Remaining chems to download
chems_to_download = list(set(chemical_list).difference(set(chemicals_downloaded)))
# Filter to non-downloaded chems
# in_chemicals = in_chemicals[in_chemicals["chems"].isin(chems_to_download)]

# Initialize master table to which we'll add each individual table
cphea_complete = pd.DataFrame()

for index, row in tqdm(in_chemicals.iterrows()):
    url = row["url"]
    out_file = f"excel_sheets/pprtv_cphea_{row['chems']}_output.xlsx"

    # Skip if this page was already downloaded
    if row["chems"] in chemicals_downloaded:
        # Check if file was downloaded/craeted successfully
        try:
            tmp = pd.read_excel(out_file)
            # If loads okay and has data (more than 0 rows), continue
            if tmp.shape[0] > 0:
                # Append to output list
                cphea_complete = pd.concat(objs=[cphea_complete, tmp])
                continue
        except ValueError:
            # Case where file did not save appropriately due to interruption
            print(f"Issue with downloaded file...for {row['chems']}")
    # Rest between requests
    time.sleep(0.25)
    # Tutorial: https://www.dataquest.io/blog/web-scraping-python-using-beautiful-soup/
    page = requests.get(url)
    # page.status_code

    soup = BeautifulSoup(page.content, "html.parser")
    ##############################################
    # Pull chemical name and casrn
    chemical = soup.find(class_="page-title")
    casrn = chemical.find_next().text.replace("CASRN", "").strip()

    # Find every <sup> tag, prepend a ^ to it
    sups = soup.find_all(name="sup")
    for sup in sups:
        # https://www.geeksforgeeks.org/change-the-tags-contents-and-replace-with-the-given-string-using-beautifulsoup/
        # new tag
        sup_new = soup.new_tag("sup")
        # input string
        sup_new.string = f"^{sup.text}"
        sup.replace_with(sup_new)
    print(f"Pulling data for {chemical.text}...")
    # Find assessment classes
    assessment_section = soup.find_all("h3", class_="pane-title")
    # Get separate divs (https://stackoverflow.com/questions/60548952/python-beautifulsoup-find-between-tags)
    # Gets the full HTML section you need (could text parse from here)
    # for assessment in assessment_section:
    #    panes = assessment.find_next("div", class_="pane-content")
    #    print(panes.prettify())
    # Extract tables in general
    # tables = []
    # for assessment in assessment_section:
    #    table = assessment.find_next("table")
    #    tables.append(pd.read_html(str(table)))
    # for c in tmp.children:
    #    print(c)

    # https://stackoverflow.com/questions/24108507/beautiful-soup-resultset-object-has-no-attribute-find-all
    # Example of how to slowly drill down into the pane while maintaining the needed elements along the way
    table_list = []
    headers_to_check = ("Estimate", "Weight")
    # Use this list to normalize output column order (and ensure columns are joined appropriately)
    out_cols = [
        "chemical",
        "casrn",
        "Last Updated",
        "Note_in_body",
        "assessment_type",
        "table_title",
        "url",
        "System",
        "RfD (mg/kg-day)",
        "Basis",
        "PoD",
        "UF",
        "Confidence",
        "Species Studied",
        "Duration",
        "Principal Study",
        "Note",
        "RfC (mg/m^3))",
        "RfC (mg/m^3)",
        "Oral Slope Factor",
        "Tumor site(s)",
        "Cancer",
        "Unit Risk Factor",
        "RfC (mg/kg-day)",
    ]

    # Process chemical page
    for assessment in assessment_section:
        # print(f"Assessment type: {assessment.text}")
        panes = assessment.find_next_siblings("div", class_="pane-content")
        for pane in panes:
            headers = pane.find_all("strong")
            for header in headers:
                # Collect caption data for headers not start with "Estimate" or "Weight"
                if not header.text.startswith(headers_to_check):
                    if len(table_list) == 0:
                        # Shouldn't reach this point, but flagging just in case
                        raise Exception(
                            "Unhandled header case...no table to add caption to"
                        )
                    # Add to last table extracted as it is found
                    tmp = table_list[-1]

                    # Loop to grab all text before linebreak, checking for None as well
                    cap_text = ""
                    ns_next_sibling = header.next_sibling
                    while ns_next_sibling is not None and ns_next_sibling.name not in (
                        "br",
                        "\n",
                    ):
                        # Handle superscripts
                        if ns_next_sibling.name in ("sup", "a"):
                            cap_text = cap_text + ns_next_sibling.text
                        else:
                            cap_text = cap_text + ns_next_sibling
                        ns_next_sibling = ns_next_sibling.next_sibling
                    tmp0 = pd.DataFrame(
                        {f"{header.text.replace(':', '').strip()}": cap_text},
                        index=[0],
                    )
                    tmp = pd.concat(objs=[tmp, tmp0], axis=1)
                    # Append processed caption column
                    table_list[-1] = tmp
                    # Move on, handled the case
                    continue
                # Extract the other headers to use as columns for the upcoming table
                # print(f"Header title: {header.text}")
                n_tag = header.find_next().find_next()
                # Handle case where there is no table, but there is text
                if n_tag.find_next().name != "table":
                    # Back up since table not found
                    n_tag = n_tag.find_previous()
                    if n_tag.name == "span":
                        # Empty dataframe so column headers match later
                        # tmp0 = pd.DataFrame(columns=table_cols)

                        # Check whether body text contains a hyperlink, and if so, extract it
                        note_text = n_tag.find_next().next_sibling
                        href = note_text.find_next().get("href")
                        if href != None:
                            note = f"{note_text.strip()} {note_text.find_next().text}: {href}"
                        else:
                            note = note_text.strip()
                        tmp = pd.DataFrame(
                            {
                                "chemical": chemical.text,
                                "casrn": casrn,
                                "Last Updated": [
                                    n_tag.text.replace("Last Updated: ", "")
                                ],
                                "Note_in_body": note,
                                "assessment_type": assessment.text,
                                "table_title": header.text,
                                "url": url,
                            }
                        )
                    else:
                        raise Exception("Unhandled header case...")
                    # Append processed dataframe
                    table_list.append(tmp.reset_index(drop=True))
                else:
                    # Found table to extract!
                    table = n_tag.find_next()
                    if table is not None:
                        # See if any cases where it's not a DataFrame list that's generated
                        # Unnesting with index 0 selection for now...
                        tmp = pd.read_html(str(table))[0]
                        # Add additional section elements
                        if not tmp.empty:
                            tmp["chemical"] = chemical.text
                            tmp["casrn"] = casrn
                            tmp["Last Updated"] = n_tag.find_previous().text.replace(
                                "Last Updated: ", ""
                            )
                            tmp["assessment_type"] = assessment.text
                            tmp["table_title"] = header.text
                            tmp["url"] = url
                            # Append processed dataframe
                            table_list.append(tmp.reset_index(drop=True))
    out = pd.concat(objs=table_list)

    # Combine the two different Note columns, which should not have any overlap
    try:
        out["Note"] = [
            i + j
            for i, j in zip(out.fillna("")["Note"], out.fillna("")["Note_in_body"])
        ]
        out = out.drop(columns="Note_in_body")
    except KeyError:
        # print("No note")
        pass
    # Sort columns if exist from out_cols list
    out = out[[x for x in out_cols if x in out.columns]]

    out.to_excel(out_file, index=False)

    cphea_complete = pd.concat(objs=[cphea_complete, out])
# Save full table and sub-tables filtered by table title
cphea_complete.to_excel("pprtv_cphea_full.xlsx", index=False)
for title in cphea_complete["table_title"].unique():
    cphea_complete.loc[cphea_complete["table_title"] == title].to_excel(
        f"pprtv_cphea_{title}.xlsx", index=False
    )
# After downloading all files, loop through them to build a comprehensive list of column names
# to add to the beginning of this file. Then redownload them and sort columns by that list.
# files_downloaded = listdir("excel_sheets")
# col_name_set = []
# for filename in files_downloaded:
#    x = pd.read_excel(f'excel_sheets/{filename}')
#    for col in x.columns:
#        if col not in col_name_set:
#            col_name_set.append(col)
